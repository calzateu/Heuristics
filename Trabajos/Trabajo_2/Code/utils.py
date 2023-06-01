
import time
import numpy as np
import matplotlib.pyplot as plt
import openpyxl
import os
import re

from VND import *
from MS_ELS import MS_ELS
from neighborhoods import check_capacity_vehicles

class Utils():
    def __init__(self) -> None:
        self.__MAX_NUM_NODOS  = 201

    def read_data(self, file_name):
        nodes       = np.zeros((self.__MAX_NUM_NODOS , 4))

        cont = 0
        with open(file_name, 'r') as file:
            for line in file:
                line         = list(map(int, line.split()))
                nodes[cont]  = line
                cont         += 1

        self.problem_information = nodes[0]
        self.nodes  = nodes[1:cont]
        self.cont   = cont - 1

        return self.problem_information, self.nodes, self.cont

    def compute_distances(self):
        n = len(self.nodes)
        dist_matrix = np.zeros((n, n))

        for i in range(n):
            for j in range(i+1, n):
                # Compute the Euclidean distance between nodes i and j
                x1, y1 = self.nodes[i][1], self.nodes[i][2]
                x2, y2 = self.nodes[j][1], self.nodes[j][2]
                dist = np.sqrt((x1 - x2)**2 + (y1 - y2)**2)

                # Update the distance matrix with the computed distance
                dist_matrix[i, j] = dist
                dist_matrix[j, i] = dist

        self.dist_matrix = dist_matrix

        return dist_matrix

    def __traveled_distance(self, path):
        distance = 0
        for i in range(len(path)-1):
            distance += self.dist_matrix[int(path[i]), int(path[i+1])]

        return distance

    def __validate_solutions(self, paths, capacity_of_vehicles, max_distance):
        total_distances_traveled = 0
        feasible_path = 0
        nodes_visited = 0

        for i in paths:
            ocupation = 0
            for j in i:
                if j == 0:
                    if ocupation > capacity_of_vehicles:
                        # if verbose:
                        #     print("Max capacity exceed", ocupation)
                        #     print(i)
                        break
                    ocupation = 0
                else:
                    ocupation += self.nodes[int(j), 3]

            distance = self.__traveled_distance(i)

            total_distances_traveled += distance
            nodes_visited += len(i)

            # if verbose and distance > method.max_distance:
            #             print("Max distance exceed", distance)
            #             print(i)

            i.append(distance)

            if distance > max_distance:
                i.append(1)
                feasible_path = 1
            else:
                i.append(0)

        # if verbose:
        #     print(paths)
        #     print("Total distance: ", total_distances_traveled)
        #     print("Num nodes visited", nodes_visited)

        return total_distances_traveled, feasible_path

    def plot_routes(self, routes):
        plt.plot(self.nodes[0][1], self.nodes[0][2], "o")
        x = self.nodes[1:,1]
        y = self.nodes[1:,2]
        plt.plot(x, y, "o")

        for i in range(len(x)):
            plt.text(x[i], y[i]+0.5, '{}'.format(i+1))

        for i in range(len(routes)):
            x = self.nodes[list(map(int, routes[i])), [1]*len(routes[i])]
            y = self.nodes[list(map(int, routes[i])), [2]*len(routes[i])]
            plt.plot(x, y, label="Vehicle: " + str(i+1))

        plt.legend()

        plt.show()

    def save_solution(self, instances, name):
        # Create a new Excel workbook
        workbook = openpyxl.Workbook()

        # Loop through each instance and create a new worksheet for it
        for instance in instances:
            # Select the active worksheet
            worksheet = workbook.create_sheet(instance['name'])

            # Write the list of lists of nodes to the worksheet row by row
            for row in instance['nodes']:
                worksheet.append(row)

        # Delete the default sheet (index 0)
        sheet_to_delete = workbook['Sheet']
        del workbook[sheet_to_delete.title]

        # Save the workbook
        workbook.save(name)

    def join_trips_vehicle(self, path_vehicle):
        path = [0]
        for trip in path_vehicle:
            path.extend(trip[1:])

        return path

    def join_trips(self, solution):
        paths = []
        for vehicle in solution:
            paths.append(self.join_trips_vehicle(vehicle))

        return paths

    def save_method(self, solution, total_time, capacity_of_vehicles, max_distance):
        paths = self.join_trips(solution)
        total_distances_traveled, feasible_path = self.__validate_solutions(paths=paths, capacity_of_vehicles=capacity_of_vehicles, max_distance=max_distance)

        paths.append([total_distances_traveled, total_time, feasible_path])

        return paths
    
    def run_method(self, method, verbose=False, graph=False):
        start = time.time()
        paths = method.search_paths()
        end = time.time()
        total_time = end - start

        if graph:
            self.plot_routes(paths)

        #total_distances_traveled, feasible_path = self.__validate_solutions(paths=paths, method=method, verbose=verbose)

        #paths.append([total_distances_traveled, total_time, feasible_path])

        return paths

    def run_instances(self, Method, name, folder_path, verbose=False, graph=False, **kwargs):
        folder_path = os.path.abspath(folder_path)
        files = os.listdir(folder_path)
        files = sorted(files)

        instances = []
        for file in files:
            if verbose:
                print(file)

            problem_information, nodes, cont = self.read_data(folder_path + "/" + file)
            dist_matrix = self.compute_distances()
            demands = nodes[:, 3].copy()


            method_instance = Method(problem_information, dist_matrix, demands, **kwargs)
            instances.append(
                {'name': file,
                 'nodes': self.run_method(method=method_instance, verbose=verbose, graph=graph)
                }
            )

        self.save_solution(instances=instances, name=name)



    def read_solutions(self, path_to_file):
        # solution = []

        # cont = 0
        # with open(path_to_file, 'r') as file:
        #     for line in file:
        #         line = list(map(int, line.split()))
        #         solution.append(line)

        #         cont += 1

        # return solution, cont


        # Leer el archivo de Excel
        excel_file = openpyxl.load_workbook(path_to_file)

        # Crear un diccionario vacío para almacenar la información de las rutas
        instances_dict = {}

        # Iterar a través de cada hoja de cálculo correspondiente a cada instancia
        for sheet_name in excel_file.sheetnames:
            # Obtener la hoja de cálculo actual
            sheet = excel_file[sheet_name]

            # Iterar a través de las filas para obtener información de las rutas
            for row in sheet.iter_rows(min_row=1, values_only=True):
                # Acceder a los valores de cada columna
                row = np.array(row)
                row = row[[True if row_1 != None else False for row_1 in row]]
                ruta = row[:-2]
                tiempo_de_servicio = row[-2]
                factibilidad = row[-1]

                # Almacenar la información de las rutas en el diccionario
                if sheet_name not in instances_dict:
                    instances_dict[sheet_name] = []
                instances_dict[sheet_name].append({
                    'ruta': [0] + ruta + [0],
                    'distancia_recorrida': tiempo_de_servicio,
                    'factibilidad': factibilidad
                })

        return instances_dict


    def __split_path(self, path_information):
        # Dividir la lista cada vez que aparezca el cero (deposito)
        sub_lists = []
        sub_list = []
        for item in path_information['ruta']:
            if item == 0:
                if sub_list != []:
                    sub_lists.append([0] + sub_list + [0])
                sub_list = []
            else:
                sub_list.append(item)

        if sub_list != []:
                    sub_lists.append(sub_list)

        return sub_lists

    def __split_information(self, solution):

        trips = []

        for path_information in solution[:-1]:
            trips.extend(self.__split_path(path_information))

        return trips

    def __initial_solution(self, solution):
        trips = self.__split_information(solution)

        traveled_distances = []

        for trip in trips:
            traveled_distances.append(self.__traveled_distance(trip))

        return trips, traveled_distances

    def VND(self, solution, neighborhoods, dist_matrix, demands, max_capacity):
        trips, traveled_distances = self.__initial_solution(solution)

        print('################# Initial solution #################')
        print(trips)
        print(traveled_distances)
        print(sum(traveled_distances))

        self.plot_routes(trips)

        j = 0
        while j < len(neighborhoods):
            new_trip, new_traveled_distances, better = neighborhoods[j](trips,
                traveled_distances, dist_matrix, demands = demands,
                max_capacity=max_capacity, num_insertions=10000, num_relocations=1000000)
            if better:
                j = 0
                trips = new_trip
                traveled_distances = new_traveled_distances

            else:
                j = j+1

        print()
        print('################# Final solution #################')
        print(trips)
        print(traveled_distances)
        print(sum(traveled_distances))

        self.plot_routes(trips)

        return trips, traveled_distances


    def apply_VND_all_instances(self, solutions, neighborhoods, name):

        folder_path = "../../mtVRP Instances"
        folder_path = os.path.abspath(folder_path)
        files = os.listdir(folder_path)
        #files = sorted(files)

        files = sorted(files, key=lambda x: (int(re.findall('\d+', x)[0]), x))

        print(files)

        print(files)

        instances = []

        for file in files:
            print(f'######### {file} #########')

            problem_information, nodes, cont = self.read_data(folder_path + "/" + file)
            dist_matrix = self.compute_distances()
            demands = nodes[:, 3].copy()

            num_vehicles = int(problem_information[1])
            max_capacity = problem_information[2]
            max_distance = problem_information[3]

            start = time.time()
            solution, traveled_distances = VND(solutions[file], neighborhoods, dist_matrix, demands, max_capacity=max_capacity, max_distance=max_distance, num_vehicles=num_vehicles)
            end = time.time()
            total_time = end - start

            instances.append(
                {'name': file,
                 'nodes': self.save_method(solution, total_time, capacity_of_vehicles=max_capacity, max_distance=max_distance)
                }
            )

        self.save_solution(instances=instances, name=name)


    def apply_MS_ELS_all_instances(self, utils, problem_information, dist_matrix,
                demands, max_capacity, max_distance, num_cars, num_insertions,
                num_relocations, num_mutations, ni, nc, nsol, neighborhoods, std, max_iterations, ruido, name):

        folder_path = "../../mtVRP Instances"
        folder_path = os.path.abspath(folder_path)
        files = os.listdir(folder_path)
        #files = sorted(files)

        files = sorted(files, key=lambda x: (int(re.findall('\d+', x)[0]), x))

        print(files)

        print(files)

        instances = []

        for file in files:
            print(f'######### {file} #########')

            problem_information, nodes, cont = utils.read_data(folder_path + "/" + file)
            dist_matrix = self.compute_distances()
            demands = nodes[:, 3].copy()

            num_vehicles = int(problem_information[1])
            max_capacity = problem_information[2]
            max_distance = problem_information[3]

            start = time.time()
            #solution, traveled_distances = MS_ELS(solutions[file], neighborhoods, dist_matrix, demands, max_capacity=max_capacity, num_vehicles=num_vehicles)
            solution, traveled_distances = MS_ELS(utils, problem_information, dist_matrix,
                demands, max_capacity, max_distance, num_cars=num_vehicles, num_insertions=num_insertions,
                num_relocations=num_relocations, num_mutations=num_mutations, ni=ni, nc=nc, nsol=nsol, neighborhoods=neighborhoods, std=std, max_iterations=max_iterations, ruido=ruido)

            end = time.time()
            total_time = end - start

            print(check_capacity_vehicles(solution, demands=demands, max_capacity=max_capacity))

            instances.append(
                {'name': file,
                 'nodes': self.save_method(solution, total_time, capacity_of_vehicles=max_capacity, max_distance=max_distance)
                }
            )

        utils.save_solution(instances=instances, name=name)


    def distance_exceed(self, traveled_distances, max_distance, num_cars):
            exceed = 0
            for i in range(len(traveled_distances)):
                if i+1 > num_cars:
                    exceed += traveled_distances[i]
                elif traveled_distances[i] > max_distance:
                    exceed += traveled_distances[i] - max_distance

            return exceed